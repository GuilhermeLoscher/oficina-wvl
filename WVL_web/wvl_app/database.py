from __future__ import annotations

import sqlite3
from contextlib import contextmanager
from decimal import Decimal
from pathlib import Path
from typing import Iterator

from .models import Client, Product, Quote, QuoteItem, QuoteStatus, format_money


class Database:
    def __init__(self, path: Path):
        self.path = path

    @contextmanager
    def connect(self) -> Iterator[sqlite3.Connection]:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        conn = sqlite3.connect(self.path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        conn.execute("PRAGMA journal_mode = WAL")
        try:
            yield conn
            conn.commit()
        finally:
            conn.close()

    def initialize(self) -> None:
        with self.connect() as conn:
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS clients (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL UNIQUE,
                    phone TEXT NOT NULL DEFAULT '',
                    vehicle TEXT NOT NULL DEFAULT '',
                    plate TEXT NOT NULL DEFAULT ''
                );

                CREATE TABLE IF NOT EXISTS products (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    description TEXT NOT NULL UNIQUE,
                    unit_price TEXT NOT NULL,
                    unit_cost TEXT NOT NULL DEFAULT '0.00'
                );

                CREATE TABLE IF NOT EXISTS quotes (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    client_id INTEGER NOT NULL,
                    created_at TEXT NOT NULL,
                    status TEXT NOT NULL,
                    subtotal TEXT NOT NULL,
                    taxes TEXT NOT NULL,
                    total TEXT NOT NULL,
                    profit TEXT NOT NULL,
                    margin_percent TEXT NOT NULL,
                    notes TEXT NOT NULL,
                    FOREIGN KEY(client_id) REFERENCES clients(id)
                );

                CREATE TABLE IF NOT EXISTS quote_items (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    quote_id INTEGER NOT NULL,
                    description TEXT NOT NULL,
                    quantity TEXT NOT NULL,
                    unit_price TEXT NOT NULL,
                    unit_cost TEXT NOT NULL,
                    discount TEXT NOT NULL,
                    subtotal TEXT NOT NULL,
                    FOREIGN KEY(quote_id) REFERENCES quotes(id) ON DELETE CASCADE
                );

                CREATE INDEX IF NOT EXISTS idx_clients_name ON clients(name);
                CREATE INDEX IF NOT EXISTS idx_products_description ON products(description);
                CREATE INDEX IF NOT EXISTS idx_quotes_created_at ON quotes(created_at);
                """
            )
            self._seed(conn)

    def _seed(self, conn: sqlite3.Connection) -> None:
        if conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]:
            return
        clients = [
            ("Cliente Exemplo", "(11) 99999-0000", "Fiat Toro", "ABC1D23"),
            ("Empresa Modelo LTDA", "(11) 3333-2222", "Hilux", "WVL2026"),
        ]
        products = [
            ("Troca de oleo premium", "280.00", "145.00"),
            ("Diagnostico eletronico completo", "180.00", "45.00"),
            ("Pastilha de freio dianteira", "420.00", "260.00"),
            ("Mao de obra especializada", "150.00", "55.00"),
        ]
        conn.executemany("INSERT OR IGNORE INTO clients(name, phone, vehicle, plate) VALUES (?, ?, ?, ?)", clients)
        conn.executemany("INSERT OR IGNORE INTO products(description, unit_price, unit_cost) VALUES (?, ?, ?)", products)

    def search_clients(self, query: str, limit: int = 20, offset: int = 0) -> list[Client]:
        with self.connect() as conn:
            rows = conn.execute(
                """
                SELECT id, name, phone, vehicle, plate
                FROM clients
                WHERE name LIKE ?
                ORDER BY name
                LIMIT ? OFFSET ?
                """,
                (f"%{query.strip()}%", limit, offset),
            ).fetchall()
        return [Client(row["id"], row["name"], row["phone"], row["vehicle"], row["plate"]) for row in rows]

    def search_products(self, query: str, limit: int = 20, offset: int = 0) -> list[Product]:
        with self.connect() as conn:
            rows = conn.execute(
                """
                SELECT id, description, unit_price, unit_cost
                FROM products
                WHERE description LIKE ?
                ORDER BY description
                LIMIT ? OFFSET ?
                """,
                (f"%{query.strip()}%", limit, offset),
            ).fetchall()
        return [Product(row["id"], row["description"], Decimal(row["unit_price"]), Decimal(row["unit_cost"])) for row in rows]

    def upsert_client(self, client: Client) -> int:
        with self.connect() as conn:
            conn.execute(
                """
                INSERT INTO clients(name, phone, vehicle, plate)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(name) DO UPDATE SET
                    phone=excluded.phone,
                    vehicle=excluded.vehicle,
                    plate=excluded.plate
                """,
                (client.name.strip(), client.phone.strip(), client.vehicle.strip(), client.plate.strip().upper()),
            )
            return int(conn.execute("SELECT id FROM clients WHERE name = ?", (client.name.strip(),)).fetchone()[0])

    def save_quote(self, quote: Quote) -> int:
        client_id = self.upsert_client(quote.client)
        with self.connect() as conn:
            cur = conn.execute(
                """
                INSERT INTO quotes(client_id, created_at, status, subtotal, taxes, total, profit, margin_percent, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    client_id,
                    quote.created_at.isoformat(),
                    quote.status.value,
                    str(quote.subtotal),
                    str(quote.taxes),
                    str(quote.total),
                    str(quote.profit),
                    str(quote.margin_percent),
                    quote.notes,
                ),
            )
            quote_id = int(cur.lastrowid)
            conn.executemany(
                """
                INSERT INTO quote_items(quote_id, description, quantity, unit_price, unit_cost, discount, subtotal)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                [
                    (
                        quote_id,
                        item.description,
                        str(item.quantity),
                        str(item.unit_price),
                        str(item.unit_cost),
                        str(item.discount),
                        str(item.subtotal),
                    )
                    for item in quote.items
                ],
            )
            self.export_monthly_workbook(conn, quote.created_at.year, quote.created_at.month)
            return quote_id

    def upsert_product(self, product: Product) -> int:
        with self.connect() as conn:
            conn.execute(
                """
                INSERT INTO products(description, unit_price, unit_cost)
                VALUES (?, ?, ?)
                ON CONFLICT(description) DO UPDATE SET
                    unit_price=excluded.unit_price,
                    unit_cost=excluded.unit_cost
                """,
                (product.description.strip(), str(product.unit_price), str(product.unit_cost)),
            )
            return int(conn.execute("SELECT id FROM products WHERE description = ?", (product.description.strip(),)).fetchone()[0])

    def list_quotes(self, limit: int = 100) -> list[dict[str, str]]:
        with self.connect() as conn:
            rows = conn.execute(
                """
                SELECT q.id, q.created_at, q.status, q.total, q.profit, q.margin_percent, c.name, c.phone, c.vehicle, c.plate
                FROM quotes q
                JOIN clients c ON c.id = q.client_id
                ORDER BY q.id DESC
                LIMIT ?
                """,
                (limit,),
            ).fetchall()
        return [dict(row) for row in rows]

    def load_quote(self, quote_id: int) -> Quote:
        with self.connect() as conn:
            row = conn.execute(
                """
                SELECT q.created_at, q.status, q.notes, c.id, c.name, c.phone, c.vehicle, c.plate
                FROM quotes q
                JOIN clients c ON c.id = q.client_id
                WHERE q.id = ?
                """,
                (quote_id,),
            ).fetchone()
            if row is None:
                raise ValueError("Orcamento nao encontrado")
            items = conn.execute(
                """
                SELECT description, quantity, unit_price, unit_cost, discount
                FROM quote_items
                WHERE quote_id = ?
                ORDER BY id
                """,
                (quote_id,),
            ).fetchall()
        quote = Quote(
            client=Client(row["id"], row["name"], row["phone"], row["vehicle"], row["plate"]),
            status=QuoteStatus(row["status"]),
            notes=row["notes"],
        )
        quote.items = [
            QuoteItem(
                item["description"],
                Decimal(item["quantity"]),
                Decimal(item["unit_price"]),
                Decimal(item["unit_cost"]),
                Decimal(item["discount"]),
            )
            for item in items
        ]
        return quote

    def clear_quotes(self) -> None:
        with self.connect() as conn:
            conn.execute("DELETE FROM quote_items")
            conn.execute("DELETE FROM quotes")

    def backup(self) -> Path:
        target = self.path.parent / "backups" / f"wvl_backup_{__import__('datetime').datetime.now().strftime('%Y%m%d_%H%M%S')}.sqlite3"
        target.parent.mkdir(parents=True, exist_ok=True)
        with self.connect() as conn:
            dest = sqlite3.connect(target)
            try:
                conn.backup(dest)
            finally:
                dest.close()
        return target

    def export_monthly_workbook(self, conn: sqlite3.Connection | None = None, year: int | None = None, month: int | None = None) -> Path:
        from datetime import date
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        today = date.today()
        year = year or today.year
        month = month or today.month
        close_conn = False
        if conn is None:
            conn = sqlite3.connect(self.path)
            conn.row_factory = sqlite3.Row
            close_conn = True
        try:
            rows = conn.execute(
                """
                SELECT q.id, q.created_at, q.status, q.subtotal, q.taxes, q.total, q.profit, q.margin_percent,
                       c.name, c.phone, c.vehicle, c.plate
                FROM quotes q
                JOIN clients c ON c.id = q.client_id
                WHERE substr(q.created_at, 1, 7) = ?
                ORDER BY q.created_at, q.id
                """,
                (f"{year:04d}-{month:02d}",),
            ).fetchall()
        finally:
            if close_conn:
                conn.close()

        out_dir = self.path.parent / "relatorios"
        out_dir.mkdir(parents=True, exist_ok=True)
        path = out_dir / f"orcamentos_{year:04d}_{month:02d}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Orcamentos"
        headers = ["ID", "Data", "Cliente", "Telefone", "Veiculo", "Placa", "Status", "Subtotal", "Impostos", "Total", "Lucro", "Margem %"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F172A")
        for row in rows:
            ws.append([row[h] for h in ["id", "created_at", "name", "phone", "vehicle", "plate", "status", "subtotal", "taxes", "total", "profit", "margin_percent"]])
        for column in ws.columns:
            width = max(len(str(cell.value or "")) for cell in column) + 2
            ws.column_dimensions[column[0].column_letter].width = min(max(width, 10), 28)
        wb.save(path)
        return path

    def dashboard_metrics(self) -> dict[str, str]:
        with self.connect() as conn:
            row = conn.execute(
                """
                SELECT
                    COALESCE(SUM(CASE WHEN status IN ('Aprovado', 'Concluido') THEN CAST(total AS REAL) ELSE 0 END), 0) AS revenue,
                    COALESCE(SUM(CASE WHEN status NOT IN ('Concluido') THEN CAST(total AS REAL) ELSE 0 END), 0) AS open_value,
                    COUNT(*) AS quote_count,
                    COALESCE(AVG(CASE WHEN status IN ('Aprovado', 'Concluido') THEN 1.0 ELSE 0.0 END), 0) AS conversion
                FROM quotes
                """
            ).fetchone()
        return {
            "revenue": format_money(Decimal(str(row["revenue"]))),
            "open_value": format_money(Decimal(str(row["open_value"]))),
            "quote_count": str(row["quote_count"]),
            "conversion": f"{Decimal(str(row['conversion'] * 100)).quantize(Decimal('0.1'))}%",
        }
